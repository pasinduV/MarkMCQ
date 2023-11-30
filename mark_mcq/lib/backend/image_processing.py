from flask import Flask, request, jsonify
import cv2
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from PIL import Image, ImageOps

app = Flask(__name__)


@app.route("/process_folder", methods=["POST"])
def process_folder():
    data = request.json  # Receive JSON data with the folder path
    project_folder_path = data["project_folder_path"]
    paper_type = data["paper_type_index"]
    correct_answers = data["answer_list"]
    project_name = data["project_name"]
    processed_image_folder = data["processed_image_folder"]
    folder_path = data["original_image_path"]

    for i in range(len(correct_answers)):
        correct_answers[i] -= 1

    if not folder_path:
        return jsonify({"error": "Folder path not provided"})

    if not os.path.exists(folder_path):
        return jsonify({"error": "Folder not found"})

    # scores = []
    make_new_excel_sheet(project_folder_path, project_name)  # create excel sheet

    for filename in os.listdir(folder_path):
        if filename.endswith(".jpg") or filename.endswith(".png"):
            image_path = os.path.join(folder_path, filename)

            # Call the existing image processing function
            if paper_type == 0:
                result = process_image_1col(
                    image_path,
                    project_folder_path,
                    correct_answers,
                    filename,
                    project_name,
                )
            elif paper_type == 1:
                result = process_image_4col(
                    image_path,
                    project_folder_path,
                    correct_answers,
                    filename,
                    project_name,
                )
            elif paper_type == 2:
                result = process_image_2col(
                    image_path,
                    project_folder_path,
                    correct_answers,
                    filename,
                    project_name,
                )

    return jsonify()


def process_image_1col(
    image_path, folder_path, correct_answers, file_name, project_name
):
    if not image_path:
        return jsonify({"error": "Image path not provided"})

    if not os.path.exists(image_path):
        return jsonify({"error": "Image file not found"})

    rotate_image(image_path, -90)

    ####parameters
    widthImage = 230
    heightImage = 800
    questionsPercol = 25
    choices = 5
    ####

    ##functions
    def FindRectangleContours(countours):
        rectangleCon = []
        for i in countours:
            area = cv2.contourArea(i)

            if area > 50:
                perimeter = cv2.arcLength(i, True)
                approximation = cv2.approxPolyDP(i, 0.02 * perimeter, True)

                if len(approximation) == 4:
                    rectangleCon.append(i)

        rectangleCon = sorted(rectangleCon, key=cv2.contourArea, reverse=True)
        # sort 4 corner polygon based on area

        return rectangleCon
        # rectCon is a list that contains all the 4 corner contours starting from the largest one

    def FindCornerPointsFunction(cont):
        perimeters = cv2.arcLength(cont, True)
        approximation = cv2.approxPolyDP(cont, 0.02 * perimeters, True)
        return approximation

    # function to re order points to identify origin and other points in biggest rectangle
    def reorderPointsFunction(Points):
        Points = Points.reshape(
            (4, 2)
        )  # change biggest contour list to 4 by 2 list/array of points
        # 4 - 4 rows or points
        # 2 - each point has 2 values (x , y)

        # add and substract to find origin points and diagonal points and other 2 corner points
        PointsNew = np.zeros((4, 1, 2), np.int32)
        add = Points.sum(1)

        PointsNew[0] = Points[np.argmin(add)]
        PointsNew[3] = Points[np.argmax(add)]
        diff = np.diff(Points, axis=1)
        PointsNew[1] = Points[np.argmin(diff)]  # [w , 0]
        PointsNew[2] = Points[np.argmax(diff)]  # [h , 0]

        return PointsNew

    def splitFunction(img):
        # split img horizontally to get rows
        rows = np.vsplit(img, 25)
        bubbleArray = []

        for r in rows:
            columns = np.hsplit(r, 5)
            for bubble in columns:
                bubbleArray.append(bubble)

        return bubbleArray

    # function to mark correct answers in the answer sheet
    def markAnswersFunction(img, IndexOfmarking, grading, answer, questions, choices):
        widthOfBox = int(img.shape[1] / choices)
        heightOfBox = int(img.shape[0] / questions)

        for x in range(25):
            ans = IndexOfmarking[x]
            centerXlocation = (ans * widthOfBox) + widthOfBox // 2
            centerYlocation = (x * heightOfBox) + heightOfBox // 2

            if grading[x] == 1:
                MarkColor = (0, 255, 0)  # if answer is correct mark it from green
            else:
                MarkColor = (0, 0, 255)  # if answer is wrong mark it from red
                MarkCorrectAnsColor = (
                    0,
                    255,
                    0,
                )  # mark correct answer of wrong one in green
                IndexOfCorrectAns = answer[x]
                cv2.circle(
                    img,
                    (
                        (IndexOfCorrectAns * widthOfBox) + widthOfBox // 2,
                        (x * heightOfBox) + heightOfBox // 2,
                    ),
                    10,
                    MarkCorrectAnsColor,
                    cv2.FILLED,
                )

            cv2.circle(
                img, (centerXlocation, centerYlocation), 10, MarkColor, cv2.FILLED
            )

        return img

    ##

    ###
    # ORIGINAL ANSWERS
    ansCol = correct_answers

    # assigning path to a variables
    img = cv2.imread(image_path)

    # IMAGE PREPROCESSING

    # resize image
    img = cv2.resize(img, (widthImage, heightImage))

    # new image
    imageContours = img.copy()
    imgBiggestContours = img.copy()

    # convert to grey scale
    imageGrey = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # add blur
    imageBlur = cv2.GaussianBlur(imageGrey, (5, 5), 1)
    # size of kernel is 5*5
    # zigma-x value = 1

    # detect edges using img scan function
    # Finding the edges of the image
    imageCanny = cv2.Canny(imageBlur, 10, 50)
    # 10 and 50 are threshold values
    # using canny edge detector we detect rectangles that we need for marking.

    # FINDING ALL CONTOURS
    # find contours - continuous curves or outlines that represent the
    # boundaries of objects or regions in an image
    countours, hierarchy = cv2.findContours(
        imageCanny, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE
    )
    # use external method - this helps to find outer edges
    # no need of approximations

    # to draw contours
    cv2.drawContours(imageContours, countours, -1, (0, 255, 0), 10)

    # FIND RECTANGLES
    rectangleContours = FindRectangleContours(countours)
    # rectangleContours is a list with all the 4 sided contours starting from largest one

    biggestContour = FindCornerPointsFunction(rectangleContours[0])
    # take the 4 corner points of biggest rectangle

    if biggestContour.size != 0:
        cv2.drawContours(
            imgBiggestContours, biggestContour, -1, (0, 255, 0), 20
        )  # draw biggest contour

        biggestContour = reorderPointsFunction(biggestContour)
        # reorder points to identify origin and 1st one etc

        point1 = np.float32(biggestContour)
        point2 = np.float32(
            [[0, 0], [widthImage, 0], [0, heightImage], [widthImage, heightImage]]
        )
        matrix = cv2.getPerspectiveTransform(point1, point2)
        image1Warp = cv2.warpPerspective(img, matrix, (widthImage, heightImage))

        # find marked answers
        # marked answer bubbles have higher pexels than normal bubbles
        # APPLY THRESOLD to find marking points
        # convert img to grey
        image1WarpGrey = cv2.cvtColor(image1Warp, cv2.COLOR_BGR2GRAY)
        # apply thresold
        image1Thresh = cv2.threshold(image1WarpGrey, 100, 250, cv2.THRESH_BINARY_INV)[1]

        # take each bubble and see how many pexel values are non-zero to find which is marked
        # devide image to a grid where each grid has one bubble
        # as here we have 5*5 bubbles split img to 25 regions

        bubbleList1 = splitFunction(image1Thresh)

        # Getting non-zero pexel values of each box
        PixelArray1 = np.zeros((questionsPercol, choices))
        countCol = 0
        countRow = 0

        for bubble1 in bubbleList1:
            totalPixels = cv2.countNonZero(bubble1)
            PixelArray1[countRow][countCol] = totalPixels
            countCol += 1
            if countCol == choices:
                countRow += 1
                countCol = 0

        # Finding index values of markings
        Index1 = []
        for x in range(0, questionsPercol):
            arr = PixelArray1[x]
            Ival1 = np.where(arr == np.amax(arr))

            Index1.append(Ival1[0][0])

        # Grade the paper
        marks1 = []
        grade = 0
        for i in range(0, questionsPercol):
            if ansCol[i] == Index1[i]:
                marks1.append(1)
                grade += 1
            else:
                marks1.append(0)
                grade += 0

        totalScore = sum(marks1)

        update_excel_sheet(
            folder_path, file_name, totalScore, project_name
        )  # append values to excel sheet

        # mark correct and wrong answers in the answer sheet
        imgResult = image1Warp.copy()
        imgResult = markAnswersFunction(
            imgResult, Index1, marks1, ansCol, questionsPercol, choices
        )

        new_image_path = os.path.join(
            folder_path, "processed images", os.path.basename(image_path)
        )
        cv2.imwrite(new_image_path, imgResult)  # save to processed images


def process_image_4col(
    image_path, folder_path, correct_answers, file_name, project_name
):
    if not image_path:
        return jsonify({"error": "Image path not provided"})

    if not os.path.exists(image_path):
        return jsonify({"error": "Image file not found"})

    ####parameters
    widthImage = 600
    heightImage = 800
    questionsPerCol = 10
    choices = 5
    ####

    # functions
    rowNumber = 10
    colNumber = 5

    def FindRectangleContours(countours):
        # filter using area
        # loop through all contours and filter area
        rectangleCon = []

        for i in countours:
            area = cv2.contourArea(i)

            if area > 500:
                perimeter = cv2.arcLength(i, True)
                approximation = cv2.approxPolyDP(i, 0.02 * perimeter, True)

                if len(approximation) == 4:
                    rectangleCon.append(i)
        rectangleCon = sorted(rectangleCon, key=cv2.contourArea, reverse=True)

        return rectangleCon

    def FindCornerPointsFunction(cont):
        perimeters = cv2.arcLength(cont, True)
        approximation = cv2.approxPolyDP(cont, 0.02 * perimeters, True)
        return approximation

    def reorderPointsFunction(Points):
        Points = Points.reshape((4, 2))
        PointsNew = np.zeros((4, 1, 2), np.int32)
        add = Points.sum(1)

        PointsNew[0] = Points[np.argmin(add)]
        PointsNew[3] = Points[np.argmax(add)]
        diff = np.diff(Points, axis=1)
        PointsNew[1] = Points[np.argmin(diff)]
        PointsNew[2] = Points[np.argmax(diff)]

        return PointsNew

    # function to devide each answer column to a grdi to extract each individual bubble.
    def splitFunction(image):
        # 1st split horizontally to get all the rows
        rows = np.vsplit(image, rowNumber)
        bubbleArray = []

        # split vertically to get individual bubbles
        for r in rows:
            columns = np.hsplit(r, colNumber)

            for bubble in columns:
                bubbleArray.append(bubble)

        return bubbleArray

    ##end of function

    ###
    # ORIGINAL ANSWERS
    ansCol1 = correct_answers[0:10]  # answers for column 1
    ansCol2 = correct_answers[10:20]  # answers for column 2
    ansCol3 = correct_answers[20:30]  # answers for column 3
    ansCol4 = correct_answers[30:40]  # answers for column 4

    # assigning path to a variables
    img = cv2.imread(image_path)

    # IMAGE PREPROCESSING STEPS

    # 1) resize image
    img = cv2.resize(img, (widthImage, heightImage))

    # new image
    # take a copy of original image to draw contours and mark biggest contours
    imageContours = img.copy()
    imgBiggestContours = img.copy()

    # convert image to grey scale
    imageGrey = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # add blur to grey scale image
    imageBlur = cv2.GaussianBlur(imageGrey, (5, 5), 1)
    # size of kernel is 5*5
    # zigma-x value = 1

    # detect edges using imgcanny function
    # Finding the edges of the image
    imageCanny = cv2.Canny(imageBlur, 10, 50)
    # 10 and 50 are threshold values
    # using canny edge detector we detect rectangles that we need for marking.

    # FINDING ALL CONTOURS
    # contours - continuous curves or outlines that represent boundaries of objects or regions in an image
    countours, hierarchy = cv2.findContours(
        imageCanny, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE
    )
    # use external method - this helps to find outer edges
    # no need of approximations

    # draw detected contours on a new copy of image
    cv2.drawContours(imageContours, countours, -1, (0, 255, 0), 10)

    # find rectangles
    rectangleContours = FindRectangleContours(countours)
    # rectangleContours is a list which stores rectangle in area descending order

    biggestContour1 = FindCornerPointsFunction(rectangleContours[0])
    biggestContour2 = FindCornerPointsFunction(rectangleContours[1])
    biggestContour3 = FindCornerPointsFunction(rectangleContours[2])
    biggestContour4 = FindCornerPointsFunction(rectangleContours[3])
    # devide 4 column answer sheet to 4 parts.
    # each column is taken as a seperate rectangle and indexed them from 1 to 4 starting from left most one

    # check weather detected rectangles actually consists of an area
    if (
        biggestContour1.size != 0
        and biggestContour2.size != 0
        and biggestContour3.size != 0
        and biggestContour4.size != 0
    ):
        # draw the 4 biggest contours in new copy of original iamge
        cv2.drawContours(imgBiggestContours, biggestContour1, -1, (0, 255, 0), 10)
        cv2.drawContours(imgBiggestContours, biggestContour2, -1, (255, 0, 0), 10)
        cv2.drawContours(imgBiggestContours, biggestContour3, -1, (225, 255, 0), 10)
        cv2.drawContours(imgBiggestContours, biggestContour4, -1, (200, 200, 200), 10)

        # reorder 4 corner points of each detected rectangle to find exact 4 corner points in correct order in anticlock wise direction
        biggestContour1 = reorderPointsFunction(biggestContour1)
        biggestContour2 = reorderPointsFunction(biggestContour2)
        biggestContour3 = reorderPointsFunction(biggestContour3)
        biggestContour4 = reorderPointsFunction(biggestContour4)

        # as image capturing can be done in several angles apply werp perspective to get bird eye view

        point1 = np.float32(biggestContour1)
        point2 = np.float32(
            [[0, 0], [widthImage, 0], [0, heightImage], [widthImage, heightImage]]
        )
        matrix1 = cv2.getPerspectiveTransform(point1, point2)
        image1Warp = cv2.warpPerspective(img, matrix1, (widthImage, heightImage))

        point3 = np.float32(biggestContour2)
        point4 = np.float32(
            [[0, 0], [widthImage, 0], [0, heightImage], [widthImage, heightImage]]
        )
        matrix2 = cv2.getPerspectiveTransform(point3, point4)
        image2Warp = cv2.warpPerspective(img, matrix2, (widthImage, heightImage))

        point5 = np.float32(biggestContour3)
        point6 = np.float32(
            [[0, 0], [widthImage, 0], [0, heightImage], [widthImage, heightImage]]
        )
        matrix3 = cv2.getPerspectiveTransform(point5, point6)
        image3Warp = cv2.warpPerspective(img, matrix3, (widthImage, heightImage))

        point7 = np.float32(biggestContour4)
        point8 = np.float32(
            [[0, 0], [widthImage, 0], [0, heightImage], [widthImage, heightImage]]
        )
        matrix4 = cv2.getPerspectiveTransform(point7, point8)
        image4Warp = cv2.warpPerspective(img, matrix4, (widthImage, heightImage))

        # convert to grey scale
        image1WarpGrey = cv2.cvtColor(image1Warp, cv2.COLOR_BGR2GRAY)
        image2WarpGrey = cv2.cvtColor(image2Warp, cv2.COLOR_BGR2GRAY)
        image3WarpGrey = cv2.cvtColor(image3Warp, cv2.COLOR_BGR2GRAY)
        image4WarpGrey = cv2.cvtColor(image4Warp, cv2.COLOR_BGR2GRAY)

        image1Thresh = cv2.threshold(image1WarpGrey, 100, 300, cv2.THRESH_BINARY_INV)[1]
        image2Thresh = cv2.threshold(image2WarpGrey, 100, 300, cv2.THRESH_BINARY_INV)[1]
        image3Thresh = cv2.threshold(image3WarpGrey, 100, 300, cv2.THRESH_BINARY_INV)[1]
        image4Thresh = cv2.threshold(image4WarpGrey, 100, 300, cv2.THRESH_BINARY_INV)[1]

        # take each individual bubbles and find pixel values of each bubble to find marked bubbles

        # 1st split each image to 50 sections
        bubbleList1 = splitFunction(image1Thresh)
        bubbleList2 = splitFunction(image2Thresh)
        bubbleList3 = splitFunction(image3Thresh)
        bubbleList4 = splitFunction(image4Thresh)

        # getting non-zero pexel values of each bubble

        # create 4 empty arrays to store marked bubbles of each answer
        PixelArray1 = np.zeros((questionsPerCol, choices))
        PixelArray2 = np.zeros((questionsPerCol, choices))
        PixelArray3 = np.zeros((questionsPerCol, choices))
        PixelArray4 = np.zeros((questionsPerCol, choices))

        # getting non-zero pexel values of each bubble

        countColumn1 = 0
        countRow1 = 0

        for bubble1 in bubbleList1:
            totalPixels1 = cv2.countNonZero(bubble1)

            # store these in array
            PixelArray1[countRow1][countColumn1] = totalPixels1
            countColumn1 += 1
            if countColumn1 == choices:
                countRow1 += 1
                countColumn1 = 0

        # for 2nd image
        countColumn2 = 0
        countRow2 = 0

        for bubble2 in bubbleList2:
            totalPixels2 = cv2.countNonZero(bubble2)
            PixelArray2[countRow2][countColumn2] = totalPixels2
            countColumn2 += 1
            if countColumn2 == choices:
                countRow2 += 1
                countColumn2 = 0

        # for 3rd image
        countColumn3 = 0
        countRow3 = 0

        for bubble3 in bubbleList3:
            totalPixels3 = cv2.countNonZero(bubble3)
            PixelArray3[countRow3][countColumn3] = totalPixels3
            countColumn3 += 1
            if countColumn3 == choices:
                countRow3 += 1
                countColumn3 = 0

        # for 4th image
        countColumn4 = 0
        countRow4 = 0

        for bubble4 in bubbleList4:
            totalPixels4 = cv2.countNonZero(bubble4)
            PixelArray4[countRow4][countColumn4] = totalPixels4
            countColumn4 += 1
            if countColumn4 == choices:
                countRow4 += 1
                countColumn4 = 0

        # finding index values of markings

        Index1 = []
        Index2 = []
        Index3 = []
        Index4 = []

        for x in range(0, questionsPerCol):
            arr1 = PixelArray1[x]
            arr2 = PixelArray2[x]
            arr3 = PixelArray3[x]
            arr4 = PixelArray4[x]

            Ival1 = np.where(arr1 == np.amax(arr1))
            Ival2 = np.where(arr2 == np.amax(arr2))
            Ival3 = np.where(arr3 == np.amax(arr3))
            Ival4 = np.where(arr4 == np.amax(arr4))

            Index1.append(Ival1[0][0])
            Index2.append(Ival2[0][0])
            Index3.append(Ival3[0][0])
            Index4.append(Ival4[0][0])

        # grading
        marks1 = []
        marks2 = []
        marks3 = []
        marks4 = []

        for x in range(0, questionsPerCol):
            if ansCol1[x] == Index1[x]:
                marks1.append(1)
            else:
                marks1.append(0)

        for y in range(0, questionsPerCol):
            if ansCol2[y] == Index2[y]:
                marks2.append(1)
            else:
                marks2.append(0)

        for z in range(0, questionsPerCol):
            if ansCol3[z] == Index3[z]:
                marks3.append(1)
            else:
                marks3.append(0)

        for k in range(0, questionsPerCol):
            if ansCol4[k] == Index4[k]:
                marks4.append(1)
            else:
                marks4.append(0)

        # final score
        score1 = sum(marks1)
        score2 = sum(marks2)
        score3 = sum(marks3)
        score4 = sum(marks4)

        TotalScore = score1 + score2 + score3 + score4
        update_excel_sheet(
            folder_path, file_name, TotalScore, project_name
        )  # append values to excel sheet


def process_image_2col(
    image_path, folder_path, correct_answers, file_name, project_name
):
    if not image_path:
        return jsonify({"error": "Image path not provided"})

    if not os.path.exists(image_path):
        return jsonify({"error": "Image file not found"})

    rotate_image(image_path, -90)
    ####parameters
    widthImage = 300
    heightImg = 700
    questionsPercol = 25
    choices = 5
    ####

    ####function###

    rowNumber = 25
    colNumber = 5

    def FindRectangleContours(countours):
        # filter using area
        # loop through all contours and filter area
        rectangleCon = []

        for i in countours:
            area = cv2.contourArea(i)

            if area > 6000:
                perimeter = cv2.arcLength(i, True)
                approximation = cv2.approxPolyDP(i, 0.02 * perimeter, True)

                if len(approximation) == 4:
                    rectangleCon.append(i)
        rectangleCon = sorted(rectangleCon, key=cv2.contourArea, reverse=True)

        return rectangleCon

    def FindCornerPointsFunction(cont):
        perimeters = cv2.arcLength(cont, True)
        approximation = cv2.approxPolyDP(cont, 0.02 * perimeters, True)
        return approximation

    def reorderPointsFunction(Points):
        Points = Points.reshape((4, 2))
        PointsNew = np.zeros((4, 1, 2), np.int32)
        add = Points.sum(1)

        PointsNew[0] = Points[np.argmin(add)]
        PointsNew[3] = Points[np.argmax(add)]
        diff = np.diff(Points, axis=1)
        PointsNew[1] = Points[np.argmin(diff)]
        PointsNew[2] = Points[np.argmax(diff)]

        return PointsNew

    def splitFunction(image):
        # 1st split horizontally to get all the rows
        rows = np.vsplit(image, rowNumber)
        bubbleArray = []

        # split vertically to get individual bubbles
        for r in rows:
            columns = np.hsplit(r, colNumber)

            for bubble in columns:
                bubbleArray.append(bubble)

        return bubbleArray

    ####end of function####

    ###
    # ORIGINAL ANSWERS
    ansCol1 = correct_answers[0:25]  # answers for column 1
    ansCol2 = correct_answers[25:]  # answers for column 2

    # assigning path to a variables
    img = cv2.imread(image_path)

    # IMAGE PREPROCESSING

    # resize image
    img = cv2.resize(img, (widthImage, heightImg))

    # new image
    # take a copy of original image to draw contours and mark biggest contours
    imageContours = img.copy()
    imgBiggestContours = img.copy()

    # convert to grey scale
    imageGrey = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # add blur
    imageBlur = cv2.GaussianBlur(imageGrey, (5, 5), 1)
    # size of kernel is 5*5
    # zigma-x value = 1

    # detect edges using img canny function
    # Finding the edges of the image
    imageCanny = cv2.Canny(imageBlur, 10, 50)
    # 10 and 50 are threshold values
    # using canny edge detector we detect rectangles that we need for marking.

    # FINDING ALL CONTOURS
    # find contours - continuous curves or outlines that represent the
    # boundaries of objects or regions in an image
    countours, hierarchy = cv2.findContours(
        imageCanny, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE
    )
    # use external method - this helps to find outer edges
    # no need of approximations

    # draw detected contours on a new copy of image
    cv2.drawContours(imageContours, countours, -1, (0, 255, 0), 1)

    # find rectangles
    rectangleContours = FindRectangleContours(countours)
    # rectangleContours is a list which stores rectangle in area descending order
    biggestContour1 = FindCornerPointsFunction(rectangleContours[0])
    biggestContour2 = FindCornerPointsFunction(rectangleContours[1])

    # check weather detected rectangles actually consists of an area
    if biggestContour1.size != 0 and biggestContour2.size != 0:
        cv2.drawContours(imgBiggestContours, biggestContour1, -1, (0, 255, 0), 10)
        cv2.drawContours(imgBiggestContours, biggestContour2, -1, (255, 0, 0), 10)

        # reorder 4 corner points of each detected rectangle to find exact 4 corner points in correct order in anticlock wise direction
        biggestContour1 = reorderPointsFunction(biggestContour1)
        biggestContour2 = reorderPointsFunction(biggestContour2)

        # as image capturing can be done in several angles apply werp perspective to get bird eye view
        # apply werp perspective for biggestContour1 to get bird eye view
        point1 = np.float32(biggestContour1)
        point2 = np.float32(
            [[0, 0], [widthImage, 0], [0, heightImg], [widthImage, heightImg]]
        )
        matrix1 = cv2.getPerspectiveTransform(point1, point2)
        image1Warp = cv2.warpPerspective(img, matrix1, (widthImage, heightImg))

        # apply werp perspective for biggestContour2 to get bird eye view
        point3 = np.float32(biggestContour2)
        point4 = np.float32(
            [[0, 0], [widthImage, 0], [0, heightImg], [widthImage, heightImg]]
        )
        matrix2 = cv2.getPerspectiveTransform(point3, point4)
        image2Warp = cv2.warpPerspective(img, matrix2, (widthImage, heightImg))

        # apply threshold
        image1WarpGrey = cv2.cvtColor(image1Warp, cv2.COLOR_BGR2GRAY)
        image2WarpGrey = cv2.cvtColor(image2Warp, cv2.COLOR_BGR2GRAY)

        image1Thresh = cv2.threshold(image1WarpGrey, 130, 300, cv2.THRESH_BINARY_INV)[1]
        image2Thresh = cv2.threshold(image2WarpGrey, 130, 300, cv2.THRESH_BINARY_INV)[1]

        # take each individual bubbles and find pixel values of each bubble to find marked bubbles

        # 1st split each image to 25 by 5 sections
        bubbleList1 = splitFunction(image1Thresh)
        bubbleList2 = splitFunction(image2Thresh)

        # getting non-zero pexel values of each box

        # create 2 empty arrays to store marked bubbles of each answer
        PixelArray1 = np.zeros((questionsPercol, choices))
        PixelArray2 = np.zeros((questionsPercol, choices))

        countColumn1 = 0
        countRow1 = 0

        for bubble1 in bubbleList1:
            totalPixels1 = cv2.countNonZero(bubble1)
            PixelArray1[countRow1][countColumn1] = totalPixels1
            countColumn1 += 1
            if countColumn1 == choices:
                countRow1 += 1
                countColumn1 = 0

        # for 2nd image
        countColumn2 = 0
        countRow2 = 0

        for bubble2 in bubbleList2:
            totalPixels2 = cv2.countNonZero(bubble2)
            PixelArray2[countRow2][countColumn2] = totalPixels2
            countColumn2 += 1
            if countColumn2 == choices:
                countRow2 += 1
                countColumn2 = 0

        # finding index values of markings

        Index1 = []
        Index2 = []

        for x in range(0, questionsPercol):
            arr1 = PixelArray1[x]
            arr2 = PixelArray2[x]

            Ival1 = np.where(arr1 == np.amax(arr1))
            Ival2 = np.where(arr2 == np.amax(arr2))

            Index1.append(Ival1[0][0])
            Index2.append(Ival2[0][0])

        # grading
        marks1 = []
        marks2 = []

        for x in range(0, questionsPercol):
            if ansCol1[x] == Index1[x]:
                marks1.append(1)
            else:
                marks1.append(0)

        for y in range(0, questionsPercol):
            if ansCol2[y] == Index2[y]:
                marks2.append(1)
            else:
                marks2.append(0)

        # final score
        score1 = sum(marks1)
        score2 = sum(marks2)
        TotalScore = score1 + score2

        update_excel_sheet(
            folder_path, file_name, TotalScore, project_name
        )  # append values to excel sheet

    cv2.waitKey(0)


def update_excel_sheet(folder_path, image, mark, project_name):
    path = folder_path + "\\" + project_name + ".xlsx"
    workbook = load_workbook(path)
    index = os.path.splitext(image)[0]
    sheet = workbook.active
    next_row = sheet.max_row + 1

    sheet[f"A{next_row}"] = str(index)
    sheet[f"B{next_row}"] = str(mark)

    workbook.save(path)


def make_new_excel_sheet(folder_path, project_name):
    workbook = Workbook()
    sheet = workbook.active
    path = folder_path + "\\" + project_name + ".xlsx"

    sheet["A1"] = "Index"
    sheet["B1"] = "marks"

    workbook.save(path)
    workbook.close()


def rotate_image(image_path, rotate_degree):
    # Open the image file
    image = Image.open(image_path)

    # Rotate the image
    rotated_image = image.rotate(
        rotate_degree, expand=True
    )  # Rotate by 90 degrees clockwise, expand=True preserves the full rotated image

    # Create a new blank canvas with dimensions of the rotated image
    new_image = Image.new(
        "RGB", rotated_image.size, (255, 255, 255)
    )  # You can adjust the background color if desired

    # Paste the rotated image onto the new canvas
    new_image.paste(rotated_image, (0, 0))

    # Save the rotated image (replacing the original image)
    new_image.save(image_path)

    # Close the image files
    image.close()
    rotated_image.close()
    new_image.close()


if __name__ == "__main__":
    app.run(debug=True)
